"""
Servicio de importación masiva de animales desde Excel (.xlsx) o CSV.

Columnas esperadas (insensible a mayúsculas y espacios):
  crotal, nombre, sexo, fecha_nacimiento, estado, madre_crotal,
  padre_crotal, lote, peso_entrada, fecha_alta, observaciones
"""
import csv
import io
from datetime import date, datetime
from typing import Optional
from dataclasses import dataclass, field

from openpyxl import load_workbook, Workbook

# Mapeo flexible de nombres de columna → clave interna
_ALIAS = {
    "crotal": "crotal",
    "numero": "crotal",
    "número": "crotal",
    "identificacion": "crotal",
    "identificación": "crotal",
    "nombre": "nombre",
    "sexo": "sexo",
    "genero": "sexo",
    "género": "sexo",
    "fecha_nacimiento": "fecha_nacimiento",
    "nacimiento": "fecha_nacimiento",
    "fecha nacimiento": "fecha_nacimiento",
    "estado": "estado",
    "situacion": "estado",
    "situación": "estado",
    "madre": "madre_crotal",
    "madre_crotal": "madre_crotal",
    "crotal_madre": "madre_crotal",
    "padre": "padre_crotal",
    "padre_crotal": "padre_crotal",
    "crotal_padre": "padre_crotal",
    "lote": "lote",
    "grupo": "lote",
    "peso": "peso_entrada",
    "peso_entrada": "peso_entrada",
    "peso entrada": "peso_entrada",
    "fecha_alta": "fecha_alta",
    "alta": "fecha_alta",
    "fecha alta": "fecha_alta",
    "observaciones": "observaciones",
    "notas": "observaciones",
    "raza": "raza",
    "crotal_madre": "madre_crotal",
    "crotal madre": "madre_crotal",
}

ESTADOS_VALIDOS = {"gestante", "lactante", "vacia", "vacía", "semental", "ternero", "recria", "recría"}
ESTADOS_NORMALIZAR = {"vacía": "vacia", "recría": "recria"}


@dataclass
class FilaImportacion:
    fila: int
    crotal: str = ""
    nombre: str = ""
    sexo: str = ""
    fecha_nacimiento: Optional[date] = None
    estado: str = "vacia"
    madre_crotal: str = ""
    padre_crotal: str = ""
    lote: str = ""
    raza: str = ""
    peso_entrada: Optional[float] = None
    fecha_alta: Optional[date] = None
    observaciones: str = ""
    errores: list = field(default_factory=list)
    avisos: list = field(default_factory=list)

    @property
    def valida(self) -> bool:
        return len(self.errores) == 0


def _normalizar_clave(clave: str) -> str:
    limpia = clave.strip().lower()
    limpia = limpia.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    limpia = " ".join(limpia.split())  # colapsa múltiples espacios
    limpia = limpia.replace(" ", "_")
    return _ALIAS.get(limpia, "")


def _parse_fecha(valor) -> Optional[date]:
    if not valor:
        return None
    if isinstance(valor, (datetime, date)):
        return valor.date() if isinstance(valor, datetime) else valor
    s = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_sexo(valor: str) -> str:
    v = valor.strip().lower()
    if v in ("h", "hembra", "f", "female", "vaca"):
        return "hembra"
    if v in ("m", "macho", "male", "toro", "ternero"):
        return "macho"
    return v


def _validar_fila(fila: FilaImportacion) -> FilaImportacion:
    if not fila.crotal:
        fila.errores.append("Crotal obligatorio")
    elif len(fila.crotal) < 5:
        fila.avisos.append(f"Crotal '{fila.crotal}' parece corto — verifica el formato SITRAN")

    if fila.sexo not in ("hembra", "macho"):
        fila.errores.append(f"Sexo inválido: '{fila.sexo}' — usa 'hembra' o 'macho'")

    estado_norm = ESTADOS_NORMALIZAR.get(fila.estado, fila.estado)
    if estado_norm not in ESTADOS_VALIDOS:
        fila.avisos.append(f"Estado '{fila.estado}' desconocido — se usará 'vacia'")
        fila.estado = "vacia"
    else:
        fila.estado = estado_norm

    if fila.fecha_alta is None:
        fila.fecha_alta = date.today()
        fila.avisos.append("Fecha alta no indicada — se usará la fecha de hoy")

    return fila


def _filas_a_registros(cabeceras: list, filas: list) -> list[FilaImportacion]:
    mapa = {i: _normalizar_clave(h) for i, h in enumerate(cabeceras)}
    resultado = []

    for num_fila, valores in enumerate(filas, start=2):
        if not any(v for v in valores if str(v).strip()):
            continue  # fila vacía

        f = FilaImportacion(fila=num_fila)
        datos = {mapa[i]: str(v).strip() if v is not None else "" for i, v in enumerate(valores) if mapa.get(i)}

        f.crotal = datos.get("crotal", "").upper()
        f.nombre = datos.get("nombre", "")
        f.sexo = _parse_sexo(datos.get("sexo", ""))
        f.fecha_nacimiento = _parse_fecha(datos.get("fecha_nacimiento"))
        f.estado = datos.get("estado", "vacia").lower().strip()
        f.madre_crotal = datos.get("madre_crotal", "").upper()
        f.padre_crotal = datos.get("padre_crotal", "").upper()
        f.lote = datos.get("lote", "")
        f.raza = datos.get("raza", "")
        f.observaciones = datos.get("observaciones", "")

        try:
            f.peso_entrada = float(datos["peso_entrada"]) if datos.get("peso_entrada") else None
        except ValueError:
            f.avisos.append("Peso no numérico — ignorado")

        f.fecha_alta = _parse_fecha(datos.get("fecha_alta"))

        resultado.append(_validar_fila(f))

    return resultado


def parsear_excel(contenido: bytes) -> list[FilaImportacion]:
    wb = load_workbook(filename=io.BytesIO(contenido), data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []
    cabeceras = [str(c) if c is not None else "" for c in filas[0]]
    return _filas_a_registros(cabeceras, filas[1:])


def parsear_csv(contenido: bytes) -> list[FilaImportacion]:
    texto = contenido.decode("utf-8-sig", errors="replace")
    dialectos = [csv.excel, csv.excel_tab]
    for dialecto in dialectos:
        try:
            reader = list(csv.reader(io.StringIO(texto), dialect=dialecto))
            if len(reader) > 1 and len(reader[0]) > 1:
                break
        except Exception:
            continue
    if not reader:
        return []
    return _filas_a_registros(reader[0], reader[1:])


def generar_plantilla_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Animales"

    cabeceras = [
        "crotal", "nombre", "sexo", "fecha_nacimiento",
        "estado", "madre_crotal", "padre_crotal", "lote",
        "peso_entrada", "fecha_alta", "observaciones",
    ]
    ws.append(cabeceras)

    # Formato cabeceras
    from openpyxl.styles import Font, PatternFill, Alignment
    verde = PatternFill("solid", fgColor="2D7A3A")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = verde
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    # Fila de ejemplo
    ws.append([
        "ES330012345678", "Vaquina", "hembra", "15/03/2020",
        "vacia", "ES330000000001", "ES330000000002", "Lote Norte",
        "380", "01/01/2024", "Vaca de ejemplo",
    ])

    # Fila de ayuda
    ws.append(["", "", "hembra / macho", "DD/MM/AAAA",
               "vacia / gestante / lactante / semental / ternero", "", "", "",
               "kg", "DD/MM/AAAA", ""])

    from openpyxl.styles import Font as F
    for cell in ws[3]:
        cell.font = F(italic=True, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
